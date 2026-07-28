"""Generate Supabase seed SQL for classes/students/qr_tokens from the school's
source Excel exports. Read-only against the source files; writes SQL + review
CSVs under supabase/seed/. Does not connect to any database.

Sources:
  - XEA4402 Keseluruhan Murid (student master, whole school, 614 students)
  - QR_Master_Merge_Result (QR token recovery, currently covers the 9 D2C
    classes / 252 students only -- the remaining students have no card yet)

Run: python scripts/generate_seed_sql.py
"""

from __future__ import annotations

import csv
import datetime as dt
import re
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
DOCS_DIR = Path(r"D:\Summit\System\docs")
SEED_DIR = REPO_ROOT / "supabase" / "seed"
SEED_DIR.mkdir(parents=True, exist_ok=True)

STUDENT_MASTER_FILE = DOCS_DIR / "XEA4402 Keseluruhan Murid as of 2026-07-19.xlsx"
QR_MERGE_FILE = DOCS_DIR / "QR_Master_Merge_Result.xlsx"

TINGKATAN_WORD_TO_NUM = {
    "SATU": 1,
    "DUA": 2,
    "TIGA": 3,
    "EMPAT": 4,
    "LIMA": 5,
    "ENAM": 6,
}


def parse_tingkatan(value: str) -> int:
    word = value.strip().upper().removeprefix("TINGKATAN").strip()
    if word not in TINGKATAN_WORD_TO_NUM:
        raise ValueError(f"Unrecognised TAHUN/TINGKATAN value: {value!r}")
    return TINGKATAN_WORD_TO_NUM[word]


def parse_ddmmyyyy(value) -> "dt.date | None":
    if value in (None, ""):
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    return dt.datetime.strptime(str(value).strip(), "%d-%m-%Y").date()


def sql_str(value) -> str:
    if value is None:
        return "NULL"
    return "'" + str(value).replace("'", "''") + "'"


def sql_date(value: "dt.date | None") -> str:
    if value is None:
        return "NULL"
    return f"'{value.isoformat()}'"


def sql_int(value) -> str:
    if value is None:
        return "NULL"
    return str(int(value))


def load_students():
    wb = openpyxl.load_workbook(STUDENT_MASTER_FILE, data_only=True, read_only=True)
    ws = wb["Worksheet"]

    classes: dict[str, dict] = {}
    students: list[dict] = []

    for row in ws.iter_rows(min_row=7, values_only=True):
        student_id = row[1]
        if student_id is None:
            continue

        class_name = row[10]
        tingkatan = row[9]
        guru_kelas = row[15]

        if class_name not in classes:
            classes[class_name] = {
                "form_level": parse_tingkatan(tingkatan),
                "homeroom_teacher_name": guru_kelas or None,
            }

        students.append(
            {
                "student_id": int(student_id),
                "full_name": row[2],
                "ic_number": str(row[3]) if row[3] is not None else None,
                "ic_type": row[4],
                "date_of_birth": parse_ddmmyyyy(row[5]),
                "study_status": row[6],
                "enrolled_at": parse_ddmmyyyy(row[7]),
                "class_joined_at": parse_ddmmyyyy(row[8]),
                "gender": row[16],
                "class_name": class_name,
            }
        )

    wb.close()
    return classes, students


def load_qr_tokens():
    wb = openpyxl.load_workbook(QR_MERGE_FILE, data_only=True, read_only=True)

    ready = []
    ws = wb["Import Ready"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        student_id = row[0]
        if student_id is None:
            continue
        ready.append(
            {
                "student_id": int(student_id),
                "token": row[5],
                "printed_class_snapshot": row[6],
            }
        )

    unmatched = []
    ws_unmatched = wb["Unmatched Cards (Review)"]
    for row in ws_unmatched.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        unmatched.append({"name_on_card": row[0], "token": row[1], "class": row[2]})

    missing = []
    ws_missing = wb["Missing Cards (Need Printing)"]
    for row in ws_missing.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        missing.append(
            {
                "student_id": row[0],
                "name": row[1],
                "class": row[2],
                "enrolled_at": row[3],
                "class_joined_at": row[4],
            }
        )

    wb.close()
    return ready, unmatched, missing


def write_review_csv(path: Path, rows: list[dict]):
    if not rows:
        return
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def main():
    classes, students = load_students()
    qr_ready, qr_unmatched, qr_missing = load_qr_tokens()

    student_ids_in_master = {s["student_id"] for s in students}
    qr_ready_matched = [q for q in qr_ready if q["student_id"] in student_ids_in_master]
    qr_ready_unmatched_id = [q for q in qr_ready if q["student_id"] not in student_ids_in_master]

    out_path = SEED_DIR / "seed_data.sql"
    with out_path.open("w", encoding="utf-8") as f:
        f.write("-- Generated by scripts/generate_seed_sql.py -- do not hand-edit, regenerate instead.\n")
        f.write(f"-- Source: {STUDENT_MASTER_FILE.name}, {QR_MERGE_FILE.name}\n")
        f.write(f"-- Students: {len(students)}, Classes: {len(classes)}, QR tokens matched: {len(qr_ready_matched)}\n\n")

        f.write("begin;\n\n")

        f.write("-- classes\n")
        for name, c in sorted(classes.items()):
            f.write(
                "insert into public.classes (name, form_level, homeroom_teacher_name) values "
                f"({sql_str(name)}, {c['form_level']}, {sql_str(c['homeroom_teacher_name'])}) "
                "on conflict (name) do update set form_level = excluded.form_level, "
                "homeroom_teacher_name = excluded.homeroom_teacher_name;\n"
            )

        f.write("\n-- students\n")
        for s in students:
            f.write(
                "insert into public.students "
                "(student_id, full_name, ic_number, ic_type, date_of_birth, gender, study_status, "
                "enrolled_at, class_joined_at, class_id) values ("
                f"{sql_int(s['student_id'])}, {sql_str(s['full_name'])}, {sql_str(s['ic_number'])}, "
                f"{sql_str(s['ic_type'])}, {sql_date(s['date_of_birth'])}, {sql_str(s['gender'])}, "
                f"{sql_str(s['study_status'])}, {sql_date(s['enrolled_at'])}, {sql_date(s['class_joined_at'])}, "
                f"(select id from public.classes where name = {sql_str(s['class_name'])})"
                ") on conflict (student_id) do update set "
                "full_name = excluded.full_name, ic_number = excluded.ic_number, ic_type = excluded.ic_type, "
                "date_of_birth = excluded.date_of_birth, gender = excluded.gender, "
                "study_status = excluded.study_status, enrolled_at = excluded.enrolled_at, "
                "class_joined_at = excluded.class_joined_at, class_id = excluded.class_id;\n"
            )

        f.write("\n-- qr_tokens (only students matched in the master list; see qr_unmatched_review.csv otherwise)\n")
        for q in qr_ready_matched:
            f.write(
                "insert into public.qr_tokens (student_id, token, printed_class_snapshot) values ("
                f"(select id from public.students where student_id = {sql_int(q['student_id'])}), "
                f"{sql_str(q['token'])}, {sql_str(q['printed_class_snapshot'])}"
                ") on conflict (token) do nothing;\n"
            )

        f.write("\ncommit;\n")

    write_review_csv(SEED_DIR / "qr_unmatched_cards_review.csv", qr_unmatched)
    write_review_csv(SEED_DIR / "qr_missing_cards_need_printing.csv", qr_missing)
    write_review_csv(
        SEED_DIR / "qr_tokens_without_master_match.csv",
        [{"student_id": q["student_id"], "token": q["token"], "printed_class_snapshot": q["printed_class_snapshot"]} for q in qr_ready_unmatched_id],
    )

    print(f"Classes: {len(classes)}")
    print(f"Students: {len(students)}")
    print(f"QR tokens matched to a master student: {len(qr_ready_matched)}")
    print(f"QR tokens with no matching student_id in master list: {len(qr_ready_unmatched_id)}")
    print(f"Students with no QR card yet (per Missing Cards sheet): {len(qr_missing)}")
    print(f"Unmatched printed cards needing manual review: {len(qr_unmatched)}")
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
